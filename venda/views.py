from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Q
from .models import Venda
from produto.models import Produto
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
import datetime

def lista_vendas(request):
    return render(request, 'venda/lista_vendas.html')

@login_required
def pedidos_aprovados(request):
    """US7: Exibir pedidos aprovados para programar entrega"""
    vendas_aprovadas = Venda.objects.filter(status='APROVADO').select_related('cliente', 'produto')
    return render(request, 'venda/pedidos_aprovados.html', {
        'vendas': vendas_aprovadas
    })

@login_required
def programar_entrega(request, venda_id):
    """US7: Programar entrega e reservar produtos"""
    venda = get_object_or_404(Venda, id=venda_id, status='APROVADO')
    
    if request.method == 'POST':
        data_entrega = request.POST.get('data_entrega')
        
        # Verificar se há estoque suficiente
        if venda.produto.estoque < venda.quantidade:
            messages.error(request, f'Estoque insuficiente! Disponível: {venda.produto.estoque}, Necessário: {venda.quantidade}')
            return redirect('pedidos_aprovados')
        
        # Programar entrega
        venda.data_entrega = data_entrega
        venda.status = 'PROGRAMADO'
        venda.gerente_responsavel = request.user
        
        # Subtrair estoque automaticamente
        produto = venda.produto
        produto.estoque -= venda.quantidade
        produto.save()
        
        venda.save()
        
        messages.success(request, f'Entrega programada para {data_entrega}. Estoque atualizado.')
        
        # Aqui você poderia adicionar notificação ao cliente e promotor
        # enviar_notificacao_programacao(venda)
        
        return redirect('pedidos_aprovados')
    
    return render(request, 'venda/programar_entrega.html', {
        'venda': venda
    })

@login_required
def entregas_hoje(request):
    """US8: Exibir entregas para hoje"""
    hoje = timezone.now().date()
    entregas = Venda.objects.filter(
        status='PROGRAMADO',
        data_entrega=hoje
    ).select_related('cliente', 'produto')
    
    return render(request, 'venda/entregas_hoje.html', {
        'entregas': entregas,
        'data_hoje': hoje
    })

@login_required
def processar_entrega(request, venda_id):
    """US8: Processar entrega"""
    venda = get_object_or_404(Venda, id=venda_id, status='PROGRAMADO')
    
    if request.method == 'POST':
        venda.status = 'PROCESSADO'
        venda.save()
        
        messages.success(request, f'Entrega processada com sucesso para {venda.cliente.nome}')
        
        # Aqui você poderia adicionar notificação ao cliente e promotor
        # enviar_notificacao_processamento(venda)
        
        return redirect('entregas_hoje')
    
    return render(request, 'venda/processar_entrega.html', {
        'venda': venda
    })

@login_required
def relatorios_produtos(request):
    """US9: Página de seleção de relatórios"""
    return render(request, 'venda/relatorios_produtos.html')

@login_required
def relatorio_estoque_baixo(request):
    """US9: Relatório de produtos com estoque baixo"""
    limite_estoque = 10  # Produtos com estoque <= 10
    produtos = Produto.objects.filter(estoque__lte=limite_estoque).order_by('estoque')
    
    return render(request, 'venda/relatorio_estoque_baixo.html', {
        'produtos': produtos,
        'limite': limite_estoque
    })

@login_required
def relatorio_promocao(request):
    """US9: Relatório de produtos em promoção (exemplo: preço < 50)"""
    produtos_promocao = Produto.objects.filter(preco__lt=50).order_by('preco')
    
    return render(request, 'venda/relatorio_promocao.html', {
        'produtos': produtos_promocao
    })

@login_required
def relatorio_por_grupo(request):
    """US9: Relatório de produtos por faixa de preço"""
    produtos_baixo = Produto.objects.filter(preco__lt=50)
    produtos_medio = Produto.objects.filter(preco__gte=50, preco__lt=100)
    produtos_alto = Produto.objects.filter(preco__gte=100)
    
    return render(request, 'venda/relatorio_por_grupo.html', {
        'produtos_baixo': produtos_baixo,
        'produtos_medio': produtos_medio,
        'produtos_alto': produtos_alto
    })

@login_required
def exportar_relatorio_pdf(request, tipo):
    """US9: Exportar relatório em PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{tipo}.pdf"'
    
    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Título
    title = Paragraph(f"Relatório de Produtos - {tipo.replace('_', ' ').title()}", styles['Title'])
    elements.append(title)
    
    # Dados baseados no tipo
    if tipo == 'estoque_baixo':
        produtos = Produto.objects.filter(estoque__lte=10)
        data = [['Nome', 'Estoque', 'Preço']]
        for produto in produtos:
            data.append([produto.nome, str(produto.estoque), f'R$ {produto.preco}'])
    elif tipo == 'promocao':
        produtos = Produto.objects.filter(preco__lt=50)
        data = [['Nome', 'Preço', 'Estoque']]
        for produto in produtos:
            data.append([produto.nome, f'R$ {produto.preco}', str(produto.estoque)])
    else:
        produtos = Produto.objects.all()
        data = [['Nome', 'Preço', 'Estoque']]
        for produto in produtos:
            data.append([produto.nome, f'R$ {produto.preco}', str(produto.estoque)])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

@login_required
def exportar_relatorio_excel(request, tipo):
    """US9: Exportar relatório em Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{tipo}.xlsx"'
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Relatório {tipo}"
    
    # Cabeçalho
    if tipo == 'estoque_baixo':
        produtos = Produto.objects.filter(estoque__lte=10)
        headers = ['Nome', 'Estoque', 'Preço']
    elif tipo == 'promocao':
        produtos = Produto.objects.filter(preco__lt=50)
        headers = ['Nome', 'Preço', 'Estoque']
    else:
        produtos = Produto.objects.all()
        headers = ['Nome', 'Preço', 'Estoque']
    
    # Escrever cabeçalho
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Escrever dados
    for row, produto in enumerate(produtos, 2):
        if tipo == 'estoque_baixo':
            ws.cell(row=row, column=1, value=produto.nome)
            ws.cell(row=row, column=2, value=produto.estoque)
            ws.cell(row=row, column=3, value=float(produto.preco))
        elif tipo == 'promocao':
            ws.cell(row=row, column=1, value=produto.nome)
            ws.cell(row=row, column=2, value=float(produto.preco))
            ws.cell(row=row, column=3, value=produto.estoque)
        else:
            ws.cell(row=row, column=1, value=produto.nome)
            ws.cell(row=row, column=2, value=float(produto.preco))
            ws.cell(row=row, column=3, value=produto.estoque)
    
    # Salvar no response
    wb.save(response)
    return response
